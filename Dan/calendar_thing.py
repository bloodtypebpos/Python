from PIL import Image, ImageDraw, ImageFont
import os
import pandas as pd
import requests
import json
import datetime
import openpyxl

apiKey = r"fcd1a20d-510f-436f-9fc4-a74fab1c0891"
dbDir = r"C:\Users\Matt\Desktop\Current Projects\Python\Dan"
fontDir = r"C:\Windows\Fonts"

imWidth = 389
imLength = 274
fontSize = 30

fnameFont = os.path.join(fontDir,"Calibrib.ttf")
font = ImageFont.truetype(fnameFont, fontSize)

theColors = ["Red","Orange","Green","Blue","Purple","Brown","Gray"]

def num2num(OldValue,OldMin,OldMax,NewMin,NewMax):
    newValue = (((float(OldValue) - float(OldMin)) * (float(NewMax) - float(NewMin))) / (float(OldMax) - float(OldMin))) + float(NewMin)
    return newValue

def FontSize(fontSize):
    font = ImageFont.truetype(fnameFont, fontSize)
    return font

def calcFontLoc1(dx,dy,text,fontSize):
    xtext = int((dx-(0.5*fontSize*len(text)))/2)
    ytext = int((fontSize)/2)
    return xtext, ytext

def makeRectangle1(img,x1,y1,x2,y2,fill,outline,thk):
    #Makes a rectangle where x,y start from top left.
    img.rectangle([(x1,y1),(x2,y2)], fill=fill)

def makeRectangle2(img,x1,y1,x2,y2,fill,outline,thk):
    #Makes a rectangle where x,y start from top left.
    img.rectangle([(x1,y1),(x2,y2)], fill=outline)
    img.rectangle([(x1+thk,y1+thk),(x2-thk,y2-thk)], fill=fill)

def makeRectangle3(img,x1,y1,x2,y2,fill,outline,thk,text,textColor,font):
    #Makes a rectangle where x,y start from top left.
    img.rectangle([(x1,y1),(x2,y2)], fill=outline)
    img.rectangle([(x1+thk,y1+thk),(x2-thk,y2-thk)], fill=fill)
    tx, ty = img.textsize(text, font=font)
    dx = x2-x1
    dy = y2-y1
    dx = int(((dx-tx)/2) + x1)
    #dx = x1 + 10
    dy = int(((dy-ty)/2) + y1)
    #dy = y1+5
    img.text((dx, dy),text,textColor,font=font)

def sideText(im,x1,y1,text,fontsize,textColor):
    tempFont = ImageFont.truetype(fnameFont, fontsize)
    tx, ty = tempFont.getsize(text)
    img_txt = Image.new("RGBA", (tx, ty), (255, 0, 0, 0))
    draw_txt = ImageDraw.Draw(img_txt)
    draw_txt.text((0,0),text,textColor,font=tempFont)
    rtext = img_txt.rotate(90, expand=1)
    txoff = 0
    tyoff = int((tx)/2)
    im.paste(rtext,(x1-txoff,y1-tyoff,x1+ty-txoff,y1+tx-tyoff),rtext)

#im = Image.new("RGB",(imWidth,imLength))
#img = ImageDraw.Draw(im)
#makeRectangle2(img, 0, 0, imWidth, imLength, "White", "Black", 2)


months = ["NA",     #0
          "JAN",    #1
          "FEB",    #2
          "MAR",    #3
          "APR",    #4
          "MAY",    #5
          "JUN",    #6
          "JUL",    #7
          "AUG",    #8
          "SEP",    #9
          "OCT",    #10
          "NOV",    #11
          "DEC"]    #12

monthDays = [0,31,29,31,30,31,30,31,31,30,31,30,31]

alphaCols = ["A",   #SUB1
             "B",   #SUB2
             "C",   #DESCRIPTION
             "D",   #START DATE
             "E"]   #END DATE


fname = os.path.join(dbDir,"input.xlsx")
wb = openpyxl.load_workbook(fname)
shts = wb.sheetnames



def makeIt(sheetname):
    im = Image.new("RGBA",(imWidth,imLength),(255, 0, 0, 0))
    img = ImageDraw.Draw(im)
    #makeRectangle2(img, 0, 0, imWidth, imLength, "White", "Black", 2)
    sht = wb[sheetname]
    maxRow = sht.max_row + 1
    SUB1s = []
    SUB2s = []
    SUB22s= []
    DESCs = []
    SUB2Dates = []
    for i in range(2,maxRow):
        SUB1 = sht[alphaCols[0]+str(i)].value
        if SUB1 not in SUB1s:
            SUB1s.append(SUB1)
    for SUB1 in SUB1s:
        SUB2row = []
        SUB22row= []
        SUB2Date= []
        for i in range(2,maxRow):
            if SUB1 == sht[alphaCols[0]+str(i)].value:
                SUB2 = sht[alphaCols[1]+str(i)].value
                SUB22row.append(SUB2)
                sDate = sht[alphaCols[3]+str(i)].value
                eDate = sht[alphaCols[4]+str(i)].value 
                SUB2Date.append([sDate,eDate])
                if SUB2 not in SUB2row:
                    SUB2row.append(SUB2)
        SUB2s.append(SUB2row)
        SUB22s.append(SUB22row)
        SUB2Dates.append(SUB2Date)
    
    startDate = datetime.datetime(2021, 1, 1, 0, 0)
    endDate = datetime.datetime(2121, 1, 1, 0, 0)
    startMonth = 13
    endMonth = -1
    theYear = 2020
    
    for SUB2Date in SUB2Dates:
        for SFDate in SUB2Date:
            sDate = SFDate[0]
            fDate = SFDate[1]
            theYear = sDate.year
            days = fDate-sDate
            theMonth = sDate.month
            if sDate.month < startMonth:
                startMonth = sDate.month
            if fDate.month > endMonth:
                endMonth = fDate.month
    
    startDate = datetime.datetime(theYear, startMonth, 1, 0, 0)
    endDate = datetime.datetime(theYear, endMonth, monthDays[endMonth], 0, 0)
    daysTotal = (endDate-startDate).days
    
    #Initial Settings             # Recommended settings:
    fontSize = 20                 # 20
    font = FontSize(fontSize)     # FontSize(fontSize)
    padLeft = 35                  # 35 
    padRight = 5                  # 5  
    padTop = 5                    # 5
    thk = 4                       # 4
    
    #Drawing the months at the top
    pads = padLeft + padRight
    monthWidth = int((imWidth-pads)/3)
    monthHeight = 30
    
    iVal = 0
    for i in range(startMonth,endMonth+1):
        theMonth = months[i]
        makeRectangle3(img, padLeft + iVal*monthWidth, padTop, padLeft + (iVal+1)*monthWidth, padTop + monthHeight, "Black", (255,0,0,0), thk, theMonth, "White", font)
        iVal = iVal + 1
    
    #Counting number of rows
    numRows = 0
    for i in range(0,len(SUB2s)):
        numRows = numRows + len(SUB2s[i])
    
    #Drawing the rows
    fontSize = 15
    font = FontSize(fontSize)
    padTop = padTop + monthHeight
    thk = 4
    pads = padLeft + padRight
    rowWidth = int(imWidth-pads)
    rowHeight = int((imLength-padTop)/numRows)
    catRows = []
    
    iVal = 0
    for i in range(0,len(SUB2s)):
        catRows.append(len(SUB2s[i]))
        for j in range(0,len(SUB2s[i])):
            for k in range(0,len(SUB22s[i])):
                if SUB22s[i][k] == SUB2s[i][j]:
                    sDate = SUB2Dates[i][k][0]
                    fDate = SUB2Dates[i][k][1]
                    dateLeft = (sDate-startDate).days
                    dateRight= (fDate-startDate).days
                    xLeft = int(num2num(dateLeft,0,daysTotal,padLeft,padLeft+rowWidth))
                    xRight = int(num2num(dateRight,0,daysTotal,padLeft,padLeft+rowWidth))
                    makeRectangle3(img, xLeft, padTop + iVal*rowHeight, xRight, padTop + (iVal+1)*rowHeight, theColors[iVal%len(theColors)], (255, 0, 0, 0), thk, SUB2s[i][j], "White", font)        
            iVal = iVal + 1
    catPad = 2
    #Drawing the Category Brackets
    iVal = 0
    for i in range(0,len(catRows)):
        makeRectangle1(img, padLeft-10, padTop + iVal*rowHeight+catPad, padLeft-8, padTop + (iVal+catRows[i])*rowHeight-catPad, "Black", "Black", 1)
        makeRectangle1(img, padLeft-10, padTop + iVal*rowHeight+catPad, padLeft, padTop + iVal*rowHeight+catPad+2, "Black", "Black", 1)
        makeRectangle1(img, padLeft-10, padTop + (iVal+catRows[i])*rowHeight-catPad, padLeft, padTop + (iVal+catRows[i])*rowHeight-catPad-2, "Black", "Black", 1)
        stupidTick = iVal + (catRows[i]/2)
        stupidTick = int(num2num(stupidTick, 0, numRows, padTop, padTop + numRows*rowHeight))
        makeRectangle1(img, padLeft-20, stupidTick+1, padLeft-10, stupidTick-1, "Black", "Black", 1)
        sideText(im, padLeft-30, stupidTick, str(SUB1s[i]), 10, "Black")
        iVal = iVal + catRows[i]
    fname = os.path.join(dbDir,sheetname + ".gif")
    im.save(fname,"GIF",transparency=0)

for sheetname in shts:
    makeIt(sheetname)
