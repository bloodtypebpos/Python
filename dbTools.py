import os
import openpyxl
import sqlite3
import time

dbDir = 'C:/Users/Sad_Matt/Desktop/Path'
fname = os.path.join(dbDir,'example.db')
conn = sqlite3.connect(fname)
c = conn.cursor()


def makeDB2(fname):  # Convert XLSX file into Database
    wb = openpyxl.load_workbook(fname, data_only=True)
    shts = wb.sheetnames
    for s in shts:
        print(s)
        rows = []
        sht = wb[s]
        check = True
        fields = []
        for row in sht.iter_rows():
            addRow = []
            if check == True:
                for cell in row:
                    addRow.append(str(cell.value))
                    fields.append(str(cell.value))
                    check = False
            else:
                for cell in row:
                    addRow.append(str(cell.value))
            rows.append(addRow)
        
        for i in range(0,len(fields)):
                if '"' in fields[i]:
                    fields[i].replace('"','``')
                if "'" in fields[i]:
                    fields[i].replace("'",'`')
        check = True         
        for row in rows:
            if check == True:
                print(row)
                print("\n")
                c.execute('DROP TABLE IF EXISTS "' + s + '"')
                query = 'CREATE TABLE "' + s + '"(id INTEGER PRIMARY KEY,'
                for r in row:
                    query = query + ' "' + r + '" TEXT,'
                query = query[:-1] + ')'
                print(query)
                print(tuple(fields))
                c.execute(query)
                check = False
            else:
                print(row)
                query = 'INSERT INTO "' + s + '"('
                for f in fields:
                    query = query + ' "' + f + '",'
                query = query[:-1] + ') VALUES ('
                for r in row:
                    query = query + '?,'
                query = query[:-1] + ')'
                c.execute(query,tuple(row))
        conn.commit()
                    

