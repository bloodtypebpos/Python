import matplotlib.pyplot as plt
import openpyxl
import os
import numpy as np

# The assignment says we can choose any problem from the list on the website. The instructor claims to have solved the first 100 problems.
# We will therefore solve a problem over 100 that looks pretty solvable as to not bore the instructor... but also because there is likely
# a good application somewhere there. We'll find the first easy problem over 100 that also has practical use
#
# Problem #102: Triangle Containment
# If considered abstractly, this problem is already widely used. It is a reduced version of the extremely useful CONVEX HULL
# Googling a solution for checking if a point is in a triangle will yield results for Barycentric Coordinate System, which is
# an interesting find but going straight to using scripts that reference the topic isn't going to help understand why it is 
# such a useful device in the first place. So, we will actually derive the tool by walking through the Convex Hull as an
# abstract concept and systematically following the formulas that will hopefully give some insight into why this works (and
# if we're lucky, how we can apply this technique outside of just triangles)
#
# The websites we will follow along with are from wolfram-alpha:
# Convex Hull: https://mathworld.wolfram.com/ConvexHull.html
# Triangle Interior: https://mathworld.wolfram.com/TriangleInterior.html

dbDir = r'C:\Users\Matt\Desktop\James'
fname = os.path.join(dbDir, 'p102_triangles.xlsx')
wb = openpyxl.load_workbook(fname)
shts = wb.sheetnames
sht = wb[shts[0]]

maxRow = sht.max_row + 1

colors = ["Red","Orange","Yellow","Green","Blue","Purple","Brown","Black","Gray"]

X1 = []
X2 = []
X3 = []
Y1 = []
Y2 = []
Y3 = []

for i in range(1,maxRow):
    X1.append(float(sht['A'+str(i)].value))
    Y1.append(float(sht['B'+str(i)].value))
    X2.append(float(sht['C'+str(i)].value))
    Y2.append(float(sht['D'+str(i)].value))
    X3.append(float(sht['E'+str(i)].value))
    Y3.append(float(sht['F'+str(i)].value))

def plotTriangle(ind,color):
    X = [X1[ind],X2[ind],X3[ind],X1[ind]]
    Y = [Y1[ind],Y2[ind],Y3[ind],Y1[ind]]
    plt.plot(X,Y,color = color)
    plt.scatter([0],[0],color="Black",marker="X")
    plt.show()

def numpyPoints(ind):
    p1 = np.array([X1[ind],Y1[ind]])
    p2 = np.array([X2[ind],Y2[ind]])
    p3 = np.array([X3[ind],Y3[ind]])
    return p1, p2, p3

def numpyVectors(ind):
    p1, p2, p3 = numpyPoints(ind)
    v0 = p1
    v1 = p2-p1
    v2 = p3-p1
    v = p0-p1
    return v0, v1, v2, v 

def aFind(v0, v1, v2, v):
    a_1 = np.linalg.det([v,v2])
    a_2 = np.linalg.det([v0,v2])
    a_3 = np.linalg.det([v1,v2])
    a = ((a_1 - a_2)/a_3)
    return a

def bFind(v0, v1, v2, v):
    b_1 = np.linalg.det([v,v1])
    b_2 = np.linalg.det([v0,v1])
    b_3 = np.linalg.det([v1,v2])
    b = -((b_1-b_2)/b_3)
    return b

def checkPoint(v0, v1, v2, v):
    outVal = 0
    a =  aFind(v0, v1, v2, v)
    b =  bFind(v0, v1, v2, v)
    if a > 0:
        if b > 0:
            if a + b < 1:
                outVal = 1
    return outVal

def checkTriangle(ind):
    p0 = np.array([0,0])
    p1, p2, p3 = numpyPoints(ind)
    v0, v1, v2, v = numpyVectors(ind)
    outVal = checkPoint(v0, v1, v2, p0)
    if outVal == 1:
        print("Origin is in triangle!")
    else:
        print("Origin is outside of triangle!")
    plotTriangle(ind, "Red")

total = 0
for i in range(0,len(X1)):
    ind = i
    p0 = np.array([0,0])
    p1, p2, p3 = numpyPoints(ind)
    v0, v1, v2, v = numpyVectors(ind)
    outVal = checkPoint(v0, v1, v2, p0)
    total = total + outVal
    
print(total)


#checkTriangle(0)
