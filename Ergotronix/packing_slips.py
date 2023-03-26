import openpyxl
import os
import datetime
import time
import pandas as pd
import docx2txt

dbDir = r'S:\PACKING-SLIP'
fnameQuery = 'fadfda'
fnameQuery = fnameQuery.replace('-', '')
print(fnameQuery)


def index_from_file_type(the_list, file_type):
    for i, s in enumerate(the_list):
        if file_type == s:
              return i
    return -2


def openpyxl_search(f_name):
    fname = os.path.join (dbDir, f_name)
    wb = openpyxl.load_workbook(fname)
    sht = wb[wb.sheetnames[0]]
    for row in sht.iter_rows():
        for cell in row:
            if cell.value and "Serial #" in str(cell.value):
                print("Gottem")
                print(cell.column)
                print(cell.row)
                print(sht.cell(cell.row, cell.column + 1).value)


if fnameQuery == "RUN":
    print("=================================================")
    print(
        "File Search -- Runs a search through a directory for files that include all or part of the given search query. Caps are not important. ")
    print("------------------------------------------------------------------------------")

    # NOTE: REMOVE COMMENT BLOCK BELOW IF YOU'RE LOOKING AT A DIFFERENT DIRECTORY
    dbDir = input("Enter a directory you'd like to search through: \n")
    dbDir = r'{}'.format(dbDir)
    fnameQuery = input("Please enter your filename query: \n")
    print("")


filetypes = []
filecode = 0
file_types = ['xlsx']
file_code = -2
f_name = "NA"

for dir in os.walk(dbDir):
    fnames = os.listdir(dir[0])
    for fname in fnames:
        filecode = fname.split(".")[-1]
        if filecode not in filetypes:
            filetypes.append(filecode)
        if fnameQuery.lower() in fname.lower():
            print(dir[0])
            print(fname)
            f_name = fname
            file_code = index_from_file_type(file_types, fname.split(".")[-1])
            print(file_code)
            print("----------------------------------------------")


if f_name != "NA" and file_code > -1:
    if file_code == 0:
        openpyxl_search(f_name)


for filetype in filetypes:
    print(filetype)



input('[ENTER] to exit or X-out of this window.')
