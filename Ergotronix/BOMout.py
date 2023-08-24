import sqlite3
import openpyxl
import os
import datetime
import texttable
from airtable import Airtable

base_key = "appW9SUX8ihLsY2YV"
api_key = "keyszzdobucJcnVXx"


alphaCols = ['Z','A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T']


years = ["2015","2016","2017","2018"]
months = ['January','February','March','April','May','June','July','August','September','October','November','December']
dbDir = "F:/PYTHON SCRIPTS/Support Files"
templateDir = "F:/PYTHON SCRIPTS/Support Files/Forms/Templates"
dwgDir = "E:/Staff/Tigrett, Matt/Drawings"


fname = os.path.join(dbDir, 'partSort.db')
conn = sqlite3.connect(fname)
c = conn.cursor()

db_columns = ['SONum', 'SODate', 'ShipByDate', 'Customer', 'ShipTo']
air_columns = ['SO No', 'SO Date', 'Ship By', 'Customer Name', 'Ship To Name']

def makeInventoryTable():
    labels = ['Item ID','Item Description','Qty on Hand']
    labelCols = []
    ##############################  INVENTORY  ##########################################
    c.execute('DROP TABLE IF EXISTS Inventory')
    conn.commit()
    c.execute('CREATE TABLE Inventory ('
              'id INTEGER PRIMARY KEY, '
              'pid TEXT, '
              'description TEXT, '
              'qty INTEGER, '
              'need INTEGER, '
              'location TEXT, '
              'Vendor TEXT, '
              'PONum TEXT, '
              'qtyOnOrder INTEGER, '
              'code TEXT, '
              'm3 INTEGER, '
              'm6 INTEGER, '
              'THICKNESS TEXT, '
              'WIDTH TEXT, '
              'LENGTH TEXT, '
              'OD TEXT, '
              'TYPE TEXT, '
              'MTL TEXT'
              ')')
    conn.commit()
    fname = os.path.join(dbDir, 'Inventory.xlsx')
    wb = openpyxl.load_workbook(fname, data_only=True)
    shtNames = wb.sheetnames
    sht = wb[shtNames[0]]
    endRow = sht.max_row-2
    
    for i in range(0,len(labels)):
        myLetter = "Z"
        for j in range(1,len(alphaCols)):
            if sht[alphaCols[j]+'1'].value == labels[i]:
                myLetter = alphaCols[j]
        labelCols.append(myLetter)
        
        
    for i in range(2,endRow):
        pid = sht[labelCols[0]+str(i)].value
        description = sht[labelCols[1]+str(i)].value
        qty = sht[labelCols[2]+str(i)].value
        if qty == None:
            qty = 0
        qty = int(qty)
        record = (pid, description, qty, 0, "", "", "", 0, "", 0, 0, "", "", "", "", "", "")
        #print(record)
        c.execute('INSERT INTO Inventory(pid, description, qty, need, location, Vendor, PONum, qtyOnOrder, code, m3, m6, THICKNESS, WIDTH, LENGTH, OD, TYPE, MTL) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)', record)
    conn.commit()
    
def makeBOMTable():
    labels = ['Item ID','Item Description','Qty Needed']
    labelCols = []
    ##############################  BILL OF MATERIALS  ##########################################
    c.execute('DROP TABLE IF EXISTS BOM')
    conn.commit()
    c.execute('CREATE TABLE BOM ('
              'id INTEGER PRIMARY KEY, '
              'assembly TEXT, '
              'assemblyDescription TEXT, '
              'pid TEXT, '
              'description TEXT, '
              'qty INTEGER'
              ')')
    conn.commit()
    fname = os.path.join(dbDir, 'BOM.xlsx')
    wb = openpyxl.load_workbook(fname, data_only=True)
    shtNames = wb.sheetnames
    sht = wb[shtNames[0]]
    endRow = sht.max_row
    
    for i in range(0,len(labels)):
        myLetter = "Z"
        for j in range(1,len(alphaCols)):
            if sht[alphaCols[j]+'1'].value == labels[i]:
                myLetter = alphaCols[j]
        labelCols.append(myLetter)     
         
    assembly = ""
    assemblyDescription = ""
    pid = ""
    description = ""
    qty = 0
    for i in range(2,endRow):
        pid = sht[labelCols[0]+str(i)].value
        #print(pid)
        if pid == None:
            pass
        elif " Total" in pid:
            pass
        else:
            if pid[0] != ".":
                assembly = sht[labelCols[0]+str(i)].value
                assemblyDescription = sht[labelCols[1]+str(i)].value
            else:
                if pid[:4] == ". . ":
                    pid = pid[4:]
                else:
                    pid = pid[2:]
                description = sht[labelCols[1]+str(i)].value
                qty = sht[labelCols[2]+str(i)].value
                if qty == None:
                    qty = 0
                qty = int(qty)
                record = (assembly, assemblyDescription, pid, description, qty)
                #print(record)
                c.execute('INSERT INTO BOM(assembly, assemblyDescription, pid, description, qty) VALUES (?,?,?,?,?)', record)
    conn.commit()

def makeOnOrderTable():
    ##############################  INVENTORY  ##########################################
    labels = ['PO No','PO Date','Vendor ID','Vendor Name','Item ID','Line Description','Qty Remaining']
    labelCols = []
    c.execute('DROP TABLE IF EXISTS onOrder')
    conn.commit()
    c.execute('CREATE TABLE onOrder ('
              'id INTEGER PRIMARY KEY, '
              'PONum TEXT, '
              'PODate TEXT, '
              'VendorID TEXT, '
              'VendorName TEXT, '
              'pid TEXT, '
              'description TEXT, '
              'qty INTEGER'
              ')')
    conn.commit()
    fname = os.path.join(dbDir, 'onOrder.xlsx')
    wb = openpyxl.load_workbook(fname, data_only=True)
    shtNames = wb.sheetnames
    sht = wb[shtNames[0]]
    endRow = sht.max_row + 1
    
    for i in range(0,len(labels)):
        myLetter = "Z"
        for j in range(1,len(alphaCols)):
            if sht[alphaCols[j]+'1'].value == labels[i]:
                myLetter = alphaCols[j]
        labelCols.append(myLetter)    
    
    for i in range(2,endRow):
        PONum = sht[labelCols[0]+str(i)].value
        PODate = sht[labelCols[1]+str(i)].value
        try:
            PODate = PODate.date()
        except:
            pass
        VendorID = sht[labelCols[2]+str(i)].value
        VendorName = sht[labelCols[3]+str(i)].value
        pid = sht[labelCols[4]+str(i)].value
        description = sht[labelCols[5]+str(i)].value
        qty = sht[labelCols[6]+str(i)].value
        if qty == None:
            qty = 0
        qty = int(qty)
        record = (PONum, PODate, VendorID, VendorName, pid, description, qty)
        #print(record)
        c.execute('INSERT INTO onOrder(PONum, PODate, VendorID, VendorName, pid, description, qty) VALUES (?,?,?,?,?,?,?)', record)
    conn.commit()

def makeOpenOrderTable():
    labels = ['SO Date','Ship By','SO No','Customer Name','Ship To Name','Ship To City','Ship To State','Item ID','Line Description','Qty Remaining']
    labelCols = []
    ##############################  INVENTORY  ##########################################
    c.execute('DROP TABLE IF EXISTS openOrders')
    conn.commit()
    c.execute('CREATE TABLE openOrders ('
              'id INTEGER PRIMARY KEY, '
              'SODate TEXT, '
              'ShipByDate TEXT, '
              'SONum TEXT, '
              'Customer TEXT, '
              'ShipTo TEXT, '
              'City TEXT, '
              'State TEXT, '
              'pid TEXT, '
              'description TEXT, '
              'qty INTEGER, '
              'month TEXT, '
              'year TEXT'
              ')')
    conn.commit()
    fname = os.path.join(dbDir, 'OpenOrdersFull.xlsx')
    wb = openpyxl.load_workbook(fname, data_only=True)
    shtNames = wb.sheetnames
    sht = wb[shtNames[0]]
    endRow = sht.max_row + 1
    
    for i in range(0,len(labels)):
        myLetter = "Z"
        for j in range(1,len(alphaCols)):
            if sht[alphaCols[j]+'1'].value == labels[i]:
                myLetter = alphaCols[j]
        labelCols.append(myLetter)
        
    for i in range(2,endRow):
        SODate = sht[labelCols[0]+str(i)].value
        ShipByDate = sht[labelCols[1]+str(i)].value
        SONum = sht[labelCols[2]+str(i)].value
        customer = sht[labelCols[3]+str(i)].value
        shipTo = sht[labelCols[4]+str(i)].value
        city = sht[labelCols[5]+str(i)].value
        state = sht[labelCols[6]+str(i)].value
        pid = sht[labelCols[7]+str(i)].value
        description = sht[labelCols[8]+str(i)].value
        qty = sht[labelCols[9]+str(i)].value
        if qty == None:
            qty = 0
        qty = int(qty)
        SODate = SODate.date()
        try:
            ShipByDate = ShipByDate.date()
        except:
            ShipByDate = ""
        SOMonth = months[SODate.month - 1]
        SOYear = str(SODate.year)
        record = (SODate, ShipByDate, SONum, customer, shipTo, city, state, pid, description, qty, SOMonth, SOYear)
        #print(record)
        c.execute('INSERT INTO openOrders(SODate, ShipByDate, SONum, Customer, ShipTo, City, State, pid, description, qty, month, year) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)', record)
    conn.commit()

def makePastOrdersTable():
    ##############################  PastOrders  ##########################################
    c.execute('DROP TABLE IF EXISTS PastOrders')
    conn.commit()
    c.execute('CREATE TABLE PastOrders ('
              'id INTEGER PRIMARY KEY, '
              'SODate TEXT, '
              'ShipByDate TEXT, '
              'SONum TEXT, '
              'Customer TEXT, '
              'ShipTo TEXT, '
              'City TEXT, '
              'State TEXT, '
              'pid TEXT, '
              'description TEXT, '
              'qty INTEGER, '
              'month TEXT, '
              'year TEXT'
              ')')
    conn.commit()
    fname = os.path.join(dbDir, 'PastOrders.xlsx')
    wb = openpyxl.load_workbook(fname, data_only=True)
    shtNames = wb.get_sheet_names()
    sht = wb.get_sheet_by_name(shtNames[0])
    endRow = sht.max_row
    for i in range(2,endRow):
        SODate = sht["A"+str(i)].value
        ShipByDate = sht["B"+str(i)].value
        SONum = sht["C"+str(i)].value
        customer = sht["D"+str(i)].value
        shipTo = sht["E"+str(i)].value
        city = sht["F"+str(i)].value
        state = sht["G"+str(i)].value
        pid = sht["H"+str(i)].value
        description = sht["I"+str(i)].value
        qty = sht["J"+str(i)].value
        if qty == None:
            qty = 0
        qty = int(qty)
        SODate = SODate.date()
        try:
            ShipByDate = ShipByDate.date()
        except:
            ShipByDate = ""
        SOMonth = months[SODate.month - 1]
        SOYear = str(SODate.year)
        record = (SODate, ShipByDate, SONum, customer, shipTo, city, state, pid, description, qty, SOMonth, SOYear)
        #print(record)
        c.execute('INSERT INTO PastOrders(SODate, ShipByDate, SONum, Customer, ShipTo, City, State, pid, description, qty, month, year) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)', record)
    conn.commit()

def makePartLocationsTable():
    ##############################  PartLocations  ##########################################
    c.execute('DROP TABLE IF EXISTS partLocations')
    conn.commit()
    c.execute('CREATE TABLE partLocations ('
              'id INTEGER PRIMARY KEY, '
              'pid TEXT, '
              'description TEXT, '
              'location TEXT, '
              'code TEXT, '
              'tier TEXT, '
              'THICKNESS TEXT, '
              'WIDTH TEXT, '
              'LENGTH TEXT, '
              'OD TEXT, '
              'TYPE TEXT, '
              'MTL TEXT'
              ')')
    conn.commit()
    fname = os.path.join(dbDir, 'PartLocations.xlsx')
    wb = openpyxl.load_workbook(fname, data_only=True)
    shtNames = wb.sheetnames
    sht = wb[shtNames[0]]
    endRow = sht.max_row + 1
    for i in range(2,endRow):
        pid = sht["A"+str(i)].value
        description = sht["B"+str(i)].value
        location = sht["C"+str(i)].value
        code = sht["D"+str(i)].value
        pthk = sht["E"+str(i)].value
        pwdt = sht["F"+str(i)].value
        plen = sht["G"+str(i)].value
        podd = sht["H"+str(i)].value
        ptyp = sht["I"+str(i)].value
        pmtl = sht["J"+str(i)].value
        record = (pid, description, location, code, pthk, pwdt, plen, podd, ptyp, pmtl)
        #print(record)
        c.execute('INSERT INTO partLocations(pid, description, location, code, THICKNESS, WIDTH, LENGTH, OD, TYPE, MTL) VALUES (?,?,?,?,?,?,?,?,?,?)', record)
    conn.commit()


def makePartNumberTable():
    ##############################  PartNumbers  ##########################################
    # THIS WAS SPECIFICALLY FOR THE SPACE PROBLEM IN 2018                                 #
    #######################################################################################
    c.execute('DROP TABLE IF EXISTS partNumbers')
    conn.commit()
    c.execute('CREATE TABLE partNumbers ('
              'id INTEGER PRIMARY KEY, '
              'pid TEXT, '
              'description TEXT, '
              'location TEXT, '
              'code TEXT, '
              'year TEXT, '
              'January INTEGER, '
              'February INTEGER, '
              'March INTEGER, '
              'April INTEGER, '
              'May INTEGER, '
              'June INTEGER, '
              'July INTEGER, '
              'August INTEGER, '
              'September INTEGER, '
              'October INTEGER, '
              'November INTEGER, '
              'December INTEGER'
              ')')
    conn.commit()
    partLocationRows = []
    rows = c.execute('SELECT * FROM partLocations')
    for row in rows:
        partLocationRows.append(row)
    for part in partLocationRows:
        pid = part[1]
        description = part[2]
        location = part[3]
        code = part[4]
        for year in years:
            record = (pid, description, location, code, year, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            #print(record)
            c.execute('INSERT INTO partNumbers(pid, description, location, code, year, January, February, March, April, May, June, July, August, September, October, November, December) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)', record)
    conn.commit()

def BOMorPart(query):
    bom = []
    parts = []
    rows = c.execute('SELECT * FROM BOM WHERE assembly = ?', (query,))
    for row in rows:
        bom.append(row)
    if len(bom) > 0:
        for b in bom:
            parts.append([b[3],b[4],b[5]])
    else:
        rows = c.execute('SELECT * FROM Inventory WHERE pid = ?', (query,))
        for row in rows:
            parts.append([row[1],row[2],1])
    return parts

def makeTroubleTable():
    c.execute('DROP TABLE IF EXISTS troubleTable')
    conn.commit()
    c.execute('CREATE TABLE troubleTable ('
              'id INTEGER PRIMARY KEY, '
              'PART TEXT, '
              'DESCRIPTION TEXT, '
              'NEED INTEGER, '
              'SONUM TEXT, '
              'DIFF INTEGER, '
              'HAVE INTEGER, '
              'hmm INTEGER, '
              'onOrder INTEGER, '
              'CODE TEXT '
              ')')
    conn.commit()
    rows = c.execute('SELECT * FROM openOrders')
    SOs = []
    orderRows = []
    for row in rows:
        orderRows.append(row)
    for row in orderRows:
        if row[3] not in SOs:
            SOs.append(row[3])
    orderRows = []
    for SO in SOs:
        boms = []
        rows = c.execute('SELECT * FROM openOrders WHERE SONum = ?', (SO,))
        for row in rows:
            boms.append([row[8],row[10]])
        for bom in boms:
            iid = bom[0]
            qty = bom[1]
            pids = BOMorPart(iid)
            for pid in pids:
                row = []
                for i in range(0,len(pid)):
                    if i == 2:
                        row.append(pid[i]*qty)
                    else:
                        row.append(pid[i])
                row.append(SO)
                row.append(0)
                row.append(0)
                row.append(0)
                row.append(0)
                row.append('NA')
                orderRows.append(row)
    for pid in orderRows:
        rows = c.execute('SELECT * FROM Inventory WHERE pid = ?', (pid[0],))
        for row in rows:
            pid[4] = row[3]-row[4]
            pid[5] = row[3]
            pid[6] = row[3] - pid[2]
            pid[7] = row[8]
            pid[8] = row[9]
            if pid[7] == "":
                pid[7] = "NA"
    #orderRows.insert(0,['PART', 'DESCRIPTION', 'NEED', 'SONUM','DIFF','HAVE','hmm'])
    SOs = []
    for row in orderRows:
        if row[4] < 0:
            SOs.append(row)
    SOs.insert(0,['PART', 'DESCRIPTION', 'NEED', 'SONUM','DIFF','HAVE','hmm', 'onOrder', 'CODE'])
    #printTable2(SOs)
    for SO in SOs:
        c.execute('INSERT INTO troubleTable(PART, DESCRIPTION, NEED, SONUM, DIFF, HAVE, hmm ,onOrder, CODE) VALUES (?,?,?,?,?,?,?,?,?)', SO)
    conn.commit()

def makeTroubleReport():
    rows = []
    SOs = c.execute('SELECT * FROM openOrders')
    for row in SOs:
        rows.append(row)
    SOs = []
    for row in rows:
        if row[3] not in SOs:
            SOs.append(row[3])
    for SO in SOs:
        print(SO)
        rows = []
        orderRows = c.execute('SELECT CUSTOMER, pid, substr(description,0,55), qty FROM openOrders WHERE SONum = ?', (SO,))
        for row in orderRows:
            rows.append(row)
        rows.insert(0,['CUSTOMER','PID','DESCRIPTION','QTY'])
        print('The order:')
        printTable2(rows)        
        print('')
        print('- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -')
        print('')
        print('The problem parts:')
        rows = []
        pids = c.execute('SELECT SONUM,PART,DESCRIPTION,HAVE,NEED,DIFF,hmm,onOrder,CODE FROM troubleTable WHERE SONUM = ? ORDER BY CODE', (SO,))
        for row in pids:
            rows.append(row)
        rows.insert(0,['SONUM','PART','DESCRIPTION','HAVE','NEED','DIFF','hmm','onOrder','CODE'])
        printTable2(rows)
        print("##############################################################################################")
        

def listOpenOrders():
    rows = []
    myRows = c.execute('SELECT SONum FROM openOrders')
    for row in myRows:
        if row[0] not in rows:
            rows.append(row[0])
    for row in rows:
        print(row)
     
def makeAllTables():
    makeInventoryTable()
    makeBOMTable()
    makePastOrdersTable()
    makePartLocationsTable()
    #makePartNumberTable()
    

def statusBar(OldMin, OldMax, NewMin, NewMax, OldValue, opt1):
    NewValue = (((OldValue - OldMin) * (NewMax - NewMin)) / (OldMax - OldMin)) + NewMin
    NewValue = int(NewValue)
    theStatus = "["
    for i in range(NewMin, NewValue):
        theStatus = theStatus + "#"
    for i in range(NewValue, NewMax):
        theStatus = theStatus + "-"
    theStatus = theStatus + "]   "
    percentVal = int((NewValue/NewMax)*100)
    theStatus = theStatus + str(percentVal) + "%   [" + str(OldValue) + "/" + str(OldMax) + "]   " + opt1
    return theStatus


#################################################################################################################################

def updatePartNumberTable():
    ##############################  PartNumbers  ##########################################
    # THIS WAS SPECIFICALLY FOR THE SPACE PROBLEM IN 2018                                 #
    #######################################################################################    
    partLocationRows = []
    rows = c.execute('SELECT * FROM partLocations')
    for row in rows:
        partLocationRows.append(row) 
    val = 0   
    for part in partLocationRows:
        print(statusBar(0,len(partLocationRows),0,100,val,part[1]))
        val = val + 1
        pid = part[1]
        bomRows = []
        rows = c.execute('SELECT * FROM BOM WHERE pid = ?', (pid,))
        for row in rows:
            bomRows.append(row)
        for year in years:
            for month in months:
                num = 0
                for bommy in bomRows:
                    bom = bommy[1]
                    rows = c.execute('SELECT * FROM PastOrders WHERE pid = ? AND month = ? AND year = ?', (bom, month, year,))
                    for row in rows:
                        num = num + 1
                c.execute('UPDATE partNumbers SET ' + month + ' = ? WHERE pid = ? AND year = ?', (str(num), pid, year,))
    conn.commit()
                

def checkFirstTier(pid):
    ##############################  PartNumbers  ##########################################
    # THIS WAS SPECIFICALLY FOR THE SPACE PROBLEM IN 2018                                 #
    #######################################################################################
    #######################################################################################
    ##################################################################
    #                        THE TIER LIST                           #
    ##################################################################
    #
    # [0] Get it out of here - These are parts we need to have but take up space
    # [1] Keep in house - We consume these rapidly
    # [2] Keep some of these in house - These parts take a LONG time to get through
    # [3] Duration Rule Investigation - These parts need to be investigated because they caught an exception looking into the duration
    
    
    # val: [0] = neither rule
    #      [1] = first rule only
    #      [2] = both rules
    #      [3] = second rule only
    rule1 = False
    rule2 = False
    endRule = False
    infoStr = ""
    for year in years:
        rows = c.execute('SELECT * FROM partNumbers WHERE pid = ? AND year = ?', (pid, year,))
        num = 0
        q1 = 0
        q2 = 0
        q3 = 0
        q4 = 0
        val = 0
        for row in rows:
            #print(row)
            for i in range(6,18):
                num = num + row[i]
                if i < 9:
                    q1 = q1 + row[i]
                elif i < 12:
                    q2 = q2 + row[i]
                elif i < 15:
                    q3 = q3 + row[i]
                else:
                    q4 = q4 + row[i]
        if num > 12:
            #print("Wahoo!   " + str(num))
            rule1 = True
            val = 1
        if q1 > 3:
            if q2 > 3:
                if q3 > 3:
                    if q4 > 3:
                        #print("WAHOO!  | " + str(q1) + " | " + str(q2) + " | " + str(q3) + " | " + str(q4))
                        rule2 = True
                        if val == 1:
                            val = 2
                        else:
                            val = 3
        if year == "2018":
            if q1 > 3:
                if q2 > 3:
                    #print("WAHOO!  | " + str(q1) + " | " + str(q2))
                    rule2 = True
                    if val == 1:
                        val = 2
                    else:
                        val = 3
        infoStr = infoStr + "[" + year + " (" + str(val) + ")] "
    #print(infoStr)
    if rule1:
        if rule2:
            endRule = True
    return [endRule, infoStr]   

            
def delete1():
    ##############################  PartNumbers  ##########################################
    # THIS WAS SPECIFICALLY FOR THE SPACE PROBLEM IN 2018                                 #
    #######################################################################################
    rows = c.execute('SELECT * FROM partLocations')
    partArray = []
    num = 0
    for row in rows:
        partArray.append(row)
    for row in partArray:
        num = num + 1
        pid = row[1]
        print(statusBar(0, len(partArray), 0, 100, num, pid))
        #print(pid)
        result = checkFirstTier(pid)
        tier = 0
        if result[0] == True:
            tier = 1
        c.execute('UPDATE partLocations SET tier = ?, info = ? WHERE pid = ?', (tier, result[1], pid,))
    conn.commit()

def delete2():
    ##############################  PartNumbers  ##########################################
    # THIS WAS SPECIFICALLY FOR THE SPACE PROBLEM IN 2018                                 #
    #######################################################################################
    col1 = []
    col2 = []
    col3 = []
    
    rows = c.execute('SELECT * FROM partLocations WHERE tier = 1')
    partRows = []
    for row in rows:
        partRows.append(row)
    
    
    #partRows = [['EIP410','EIP410']]
    for part in partRows:
        rule1 = True
        pid = part[1]
        try:
            #print(pid)
            purchaseRows = []
            rows = c.execute('SELECT * FROM pastPurchases WHERE ItemID = ?', (pid,))
            for row in rows:
                purchaseRows.append(row)
            avgPurchaseQty = 0
            num = 0
            for purchase in purchaseRows:
                num = num + 1
                avgPurchaseQty = avgPurchaseQty + int(purchase[5])
            avgPurchaseQty = avgPurchaseQty/num
            #print(avgPurchaseQty)
            consumeRows = []
            rows = c.execute('SELECT * FROM partsMonths WHERE PART = ?', (pid,))
            for row in rows:
                consumeRows.append(row)
            for i in range(0,12):
                m1 = i%12
                m2 = (i+1)%12
                m3 = (i+2)%12
                m4 = (i+3)%12
                m5 = (i+3)%12
                m6 = (i+3)%12
                #theMonths = [months[m1]]
                #theMonths = [months[m1], months[m2]]
                #theMonths = [months[m1], months[m2], months[m3]]
                #theMonths = [months[m1], months[m2], months[m3], months[m4]]
                #theMonths = [months[m1], months[m2], months[m3], months[m4], months[m5]]
                theMonths = [months[m1], months[m2], months[m3], months[m4], months[m5], months[m6]]
                num = 0
                for mnth in theMonths:
                    for cRow in consumeRows:
                        if cRow[5] == mnth:
                            num = num + cRow[4]
                #print(str(num) + " <-- consumed in 3 months from " + theMonths[0])
                #print(str(avgPurchaseQty))
                if num > avgPurchaseQty:
                    rule1 = False
            if rule1:
                c.execute('UPDATE partLocations SET tier = ? WHERE pid = ?', (2, pid,))
                col2.append([part[1],part[2]])
            else:
                col1.append([part[1],part[2]])
        except:
            col3.append([part[1],part[2]])
            c.execute('UPDATE partLocations SET tier = ? WHERE pid = ?', (3, pid,))
    conn.commit()
    print("start")
    for col in col1:
        print(col)
    print("-------------------------")
    for col in col2:
        print(col)
    print("-------------------------")
    for col in col3:
        print(col)

def delete3():
    fname = os.path.join(dbDir, 'PartLocations.xlsx')
    wb = openpyxl.load_workbook(fname, data_only=True)
    shtNames = wb.get_sheet_names()
    sht = wb.get_sheet_by_name(shtNames[0])
    endRow = sht.max_row + 1
    for i in range(2,endRow):
        sht["C"+str(i)].value = "1-" + sht["C"+str(i)].value
    wb.save(fname)

def makeInventoryReport(withAirtable):
    #########################################################################
    #                PHASE 0: Initialize all tables
    #########################################################################    
    print("Making Inventory table")
    makeInventoryTable()
    print("Making BOM table")
    makeBOMTable()
    print("Making Open Sales Orders table")
    makeOnOrderTable()
    print("Making Purchased Items table")
    makeOpenOrderTable()    
    print("Making Part Locations table")
    makePartLocationsTable()
    #########################################################################
    #                PHASE 1: Get number of items needed
    #########################################################################
    print("Claiming items for standard products")
    SORows = []
    rows = c.execute('SELECT * FROM openOrders')
    for row in rows:
        SORows.append(row)
    for SORow in SORows:
        try:
            # SORow[8] -> Item ID
            # SORow[10] -> qty left on order
            #print([SORow[8],SORow[10]])
            scalar = SORow[10]
            BOM = []
            rows = c.execute('SELECT * FROM BOM WHERE assembly = ?', (SORow[8],))
            for row in rows:
                BOM.append(row)
            if len(BOM) < 1:
                BOM = []
                rows = c.execute('SELECT * FROM Inventory WHERE pid = ?', (SORow[8],))
                for row in rows:
                    BOM.append(row)
                for b in BOM:
                    #print(b)
                    num = scalar
                    val = b[4] + num
                    c.execute('UPDATE Inventory SET need = ? WHERE pid = ?', (str(val), b[1],))
                    
            else:
                for b in BOM:
                    #print(b)
                    num = scalar*b[5]
                    val = 0
                    rows = c.execute('SELECT * FROM Inventory WHERE pid = ?', (b[3],))
                    for row in rows:
                        val = row[4]
                    val = val + num
                    c.execute('UPDATE Inventory SET need = ? WHERE pid = ?', (str(val), b[3],))
        except:
            print(SORow)
    conn.commit()
    #########################################################################
    #                PHASE 2: Find out if it's on order and populate db
    #########################################################################
    print("Checking if any items are already on order")
    PORows = []
    rows = c.execute('SELECT * FROM onOrder')
    for row in rows:
        PORows.append(row)
    for PORow in PORows:
        #print([PORow[1],PORow[4],PORow[5],PORow[7]])
        try:
            c.execute('UPDATE Inventory SET PONum = ?, Vendor = ?, qtyOnOrder = ? WHERE pid = ?', (PORow[1], PORow[4],PORow[7],PORow[5],))
        except:
            pass
    conn.commit()
    #########################################################################
    #                PHASE 3: populate locations
    #########################################################################
    print('Locating parts')
    locRows = []
    rows = c.execute('SELECT pid, location, code, THICKNESS, WIDTH, LENGTH, OD, TYPE, MTL FROM partLocations')
    for row in rows:
        locRows.append(row)
    for locRow in locRows:
        #print([locRow[1],locRow[3],locRow[4]])
        try:
            c.execute('UPDATE Inventory SET location = ?, code = ?, THICKNESS = ?, WIDTH = ?, LENGTH = ?, OD = ?, TYPE = ?, MTL = ? WHERE pid = ?', (locRow[1],locRow[2],locRow[3],locRow[4],locRow[5],locRow[6],locRow[7],locRow[8],locRow[0]))
        except:
            pass
    conn.commit()   
    #########################################################################
    #                PHASE 4: Airtable Integration
    #########################################################################
    if withAirtable == True:
        print('Integrating Airtable projects')  
        airtableA()
    else:
        print('Airtable not included')
    #########################################################################
    #                PHASE 5: Order information... should we order?
    #########################################################################
    print('Performing Part Projection')
    todayMonth = datetime.datetime.today().month - 1
    theMonths = []
    for i in range(0,6):
        num = (i + todayMonth)%12
        theMonths.append(months[num])
    #print(theMonths) 
    invList = []
    rows = c.execute('SELECT pid FROM Inventory')
    for row in rows:
        invList.append(row[0])
    for pid in invList:
        m3 = 0
        m6 = 0
        for i in range(0,6):
            val = 0
            rows = c.execute('SELECT * FROM partsMonths WHERE PART = ? AND MONTH = ?', (pid, theMonths[i],))
            for row in rows:
                val = int(row[4])
            m6 = m6 + val
            if i < 3:
                m3 = m3 + val
        c.execute('UPDATE Inventory SET m3 = ?, m6 = ? WHERE pid = ?', (m3, m6, pid,))
    conn.commit()
    #########################################################################
    #                PHASE 6: Report Output
    #########################################################################
    print('Generating Final Report')
    num = 2
    fname = os.path.join(dbDir, 'BomTemplate2.xlsx')
    wb = openpyxl.load_workbook(fname, data_only=True)
    shtNames = wb.sheetnames
    sht = wb[shtNames[0]]
    #################### CHECKS IF A DWG IS IN MATT'S STAFF FOLDER#############
    mRows = []
    rows = c.execute('SELECT pid FROM Inventory WHERE code = "Machine" and need > 0')
    for row in rows:
        mRows.append(row[0])
    for part in mRows:
        for path, subdirs, files in os.walk(dwgDir):
            for name in files:
                if part == name.upper().split(".PDF")[0]:
                    c.execute('UPDATE Inventory SET PONum = ? WHERE pid = ?', ("DWG", part,))
                    break
    conn.commit()
    ###########################################################################
    invRows = []
    rows = c.execute('SELECT * FROM Inventory WHERE need > 0')
    for row in rows:
        invRows.append(row)
    for row in invRows:
        val = 0
        #print([row[1],row[2],row[3],row[4],"diff",row[5],row[7],row[6],row[8],row[10],row[11],row[9]])
        val = row[3]-row[4]
        sht["A"+str(num)].value = row[1]    # PID
        sht["B"+str(num)].value = row[2]    # Description
        sht["C"+str(num)].value = row[3]    # Have
        sht["D"+str(num)].value = row[4]    # Need
        sht["E"+str(num)].value = val       # Diff
        sht["F"+str(num)].value = row[5]    # Location
        sht["G"+str(num)].value = row[7]    # PO Number
        sht["H"+str(num)].value = row[6]    # Vendor
        sht["I"+str(num)].value = row[8]    # Qty On Order
        sht["J"+str(num)].value = row[10]   # 3 month
        sht["K"+str(num)].value = row[11]   # 6 month
        sht["L"+str(num)].value = row[9]    # Code
        num = num + 1
    fname = os.path.join(dbDir, 'BOMout.xlsx')
    wb.save(fname)
    print('Complete')
        
def airtableA():
    #########################################################################
    #                Updates Inventory DB Table with machining
    #                and engineering inventory items
    #########################################################################
    
    print("############################################################################################")
    print("Getting list of parts from airtable...")
    print(" - - - - - - - - - - - - - - - - - - - -")
    airtable = Airtable(base_key, 'Procurement and Fabrication', api_key)
    #records = airtable.get_all(view = 'Production', formula = "FIND('Machine', {Code})=1")    #Original
    records = airtable.get_all(view = 'Production', formula = "FIND('Inventory', {Code})=1")
    airParts = []
    invParts = []
    for record in records:
        if "Part" in record['fields'].keys():
            print(record)
            invParts.append(record)
            #if "Status" in record['fields'].keys():
            #    if record['fields']['Status'] == "Inventory":
            #        invParts.append(record)
            #        print(record)
            #    else:
            #        airParts.append(record)
            #else:
            #    airParts.append(record)
    for record in airParts:
        c.execute('UPDATE Inventory SET Vendor = ? WHERE pid = ?', ("Machine Shop", record['fields']['Part'],))
    for record in invParts:
        #print([record['fields']['Part'],record['fields']['Qty']])
        rows = c.execute('SELECT * FROM Inventory WHERE pid = ?', (record['fields']['Part'],))
        for row in rows:
            #print([row[1],row[4]])
            val = row[4] + record['fields']['Qty']
            c.execute('UPDATE Inventory SET need = ? WHERE pid = ?', (val, record['fields']['Part'],))
    conn.commit()
    print("############################################################################################")
    
def updateLeadTimes():
    ###########################################################
    #        This program is a mess... don't use it... just
    #        be careful and use as reference when making
    #        a new function that won't overwrite this one.
    ###########################################################
    c.execute('DROP TABLE IF EXISTS leadTimes')
    conn.commit()
    c.execute('CREATE TABLE leadTimes ('
              'id INTEGER PRIMARY KEY, '
              'pid TEXT, '
              'PONum TEXT, '
              'PODate TEXT, '
              'IVDate TEXT, '
              'leadTimeDay Integer'
              ')')
    c.execute('DROP TABLE IF EXISTS temp2')
    conn.commit()
    c.execute('CREATE TABLE temp2 ('
              'id INTEGER PRIMARY KEY, '
              'PONum TEXT, '
              'IVDate TEXT'
              ')')
    conn.commit()
    fname = os.path.join(dbDir, 'closedPurchases.xlsx')
    wb1 = openpyxl.load_workbook(fname, data_only=True)
    shtNames = wb1.get_sheet_names()
    sht1 = wb1.get_sheet_by_name(shtNames[0])
    endRow1 = sht1.max_row+1
    fname = os.path.join(dbDir, 'closedInvoices.xlsx')
    wb2 = openpyxl.load_workbook(fname, data_only=True)

    shtNames = wb2.get_sheet_names()
    sht2 = wb2.get_sheet_by_name(shtNames[0])
    endRow2 = sht2.max_row+1
    for i in range(2,endRow1):
        pid = sht1["F"+str(i)].value
        PODate = sht1["B"+str(i)].value.date()
        PONum = sht1["A"+str(i)].value
        record = (pid, PONum, PODate, "", 0)
        c.execute('INSERT INTO leadTimes(pid, PONum, PODate, IVDate, leadTimeDay) VALUES (?,?,?,?,?)', record)
    conn.commit()
    for i in range(2,endRow2):

        IVDate = sht2["F"+str(i)].value.date()
        PONum = sht2["I"+str(i)].value
        record = (PONum, IVDate)
        c.execute('INSERT INTO temp2(PONum, IVDate) VALUES (?,?)', record)   
    conn.commit()
    IVRows = []
    rows = c.execute('SELECT * FROM temp2')
    for row in rows:
        IVRows.append(row)
    for row in IVRows:
        c.execute('UPDATE leadTimes SET IVDate = ? WHERE PONum = ?', (row[2], row[1],))
    conn.commit()
    PORows = []
    rows = c.execute('SELECT * FROM leadTimes')
    for row in rows:
        PORows.append(row)
    for row in PORows:
        try:
            daysDiff = datetime.datetime.strptime(row[4], "%Y-%m-%d").date()-datetime.datetime.strptime(row[3], "%Y-%m-%d").date()
            c.execute('UPDATE leadTimes SET leadTimeDay = ? WHERE id = ?', (daysDiff.days, row[0]))
        except:
            pass
    conn.commit()
    c.execute('DROP TABLE IF EXISTS temp2')
    conn.commit()
    ###########################################################
    #        This inserted from an OLD file...
    ###########################################################
    fname = os.path.join(dbDir, 'leadtimeList.xlsx')
    wb1 = openpyxl.load_workbook(fname, data_only=True)
    shtNames = wb1.get_sheet_names()
    sht1 = wb1.get_sheet_by_name(shtNames[0])
    endRow1 = sht1.max_row+1
    for i in range(2,endRow1):
        try:
            pid = sht1["A"+str(i)].value
            PODate = sht1["C"+str(i)].value.date()
            PONum = sht1["H"+str(i)].value
            IVDate = sht1["D"+str(i)].value.date()
            daysDiff = sht1["E"+str(i)].value
            record = (pid, PONum, PODate, IVDate, int(daysDiff))
            c.execute('INSERT INTO leadTimes(pid, PONum, PODate, IVDate, leadTimeDay) VALUES (?,?,?,?,?)', record)
        except:
            pass
    conn.commit()    
    
def checkLeadTime(pid):
    leadTime = 0
    try:
        PORows = []
        pidRows = []
        rows = c.execute('SELECT * FROM leadTimes WHERE pid = ? and leadTimeDay !=0', (pid,))
        for row in rows:
            if row[2] not in PORows:
                PORows.append(row[2])
                pidRows.append(row)
            else:
                pass
        num = 0
        for row in pidRows:
            num = num + row[5]
        num = num/len(pidRows)
        leadTime = int(num)
    except:
        pass
    return leadTime

def checkAvailability(pidArray):
    bomRows = []
    rows = c.execute('SELECT * FROM BOM WHERE assembly = ?', (pidArray[0],))
    for row in rows:
        bomRows.append(row)
    print(len(bomRows))
    if len(bomRows) < 1:
        bomRows = []
        rows = c.execute('SELECT * FROM Inventory WHERE pid = ?',(pidArray[0],))
        for row in rows:
            print(row)
    else:
        checkRows = []
        for row in bomRows:
            pid = row[3]
            qty = pidArray[1]*row[5]
            checkRows.append([pid,qty,0])
        for checkRow in checkRows:
            rows = c.execute('SELECT pid,description,qty,need FROM Inventory WHERE pid = ?', (checkRow[0],))
            badRows = []
            for row in rows:
                badRows.append(row)
            for row in badRows:
                qtyDiff = row[2]-row[3]
                if qtyDiff < 0:
                    lt = checkLeadTime(row[0])
                    print("pid: " + str(row[0]) + " | qty on hand: " + str(row[2]) + " | qty need: " + str(row[3]) + " | lead time: " + str(lt) + "    | " + str(row[1]))

def makePackingSlips():
    c.execute('DROP TABLE IF EXISTS shipping')
    conn.commit()
    c.execute('CREATE TABLE shipping ('
              'id INTEGER PRIMARY KEY, '
              'SONum TEXT, '
              'Customer1 TEXT, '
              'address11 TEXT, '
              'address12 TEXT, '
              'city1 TEXT, '
              'state1 TEXT, '
              'zip1 TEXT, '
              'country1 TEXT, '
              'Customer2 TEXT, '
              'address21 TEXT, '
              'address22 TEXT, '
              'city2 TEXT, '
              'state2 TEXT, '
              'zip2 TEXT, '
              'country2 TEXT, '
              'pid TEXT, '
              'description TEXT, '
              'qty INTEGER'
              ')')
    conn.commit()
    fname = os.path.join(dbDir, 'Shipping.xlsx')
    wb1 = openpyxl.load_workbook(fname, data_only=True)
    shtNames = wb1.get_sheet_names()
    sht1 = wb1.get_sheet_by_name(shtNames[0])
    endRow1 = sht1.max_row + 1
    num = 2
    SONums = []
    for i in range(2,endRow1):
        SONum = sht1["C"+str(i)].value
        if SONum not in SONums:
            SONums.append(SONum)
        Customer1 = sht1["D"+str(i)].value
        address11 = sht1["P"+str(i)].value
        address12 = sht1["Q"+str(i)].value
        city1 = sht1["R"+str(i)].value
        state1 = sht1["S"+str(i)].value
        zip1 = sht1["T"+str(i)].value
        country1 = sht1["U"+str(i)].value
        Customer2 = sht1["E"+str(i)].value
        address21 = sht1["F"+str(i)].value
        address22 = sht1["V"+str(i)].value
        city2 = sht1["G"+str(i)].value
        state2 = sht1["H"+str(i)].value
        zip2 = sht1["W"+str(i)].value
        country2 = sht1["X"+str(i)].value
        pid = sht1["I"+str(i)].value
        description = sht1["J"+str(i)].value
        qty = sht1["M"+str(i)].value
        record = (num,SONum,Customer1,address11,address12,city1,state1,zip1,country1,Customer2,address21,address22,city2,state2,zip2,country2, pid, description, qty)
                #print(record)
        c.execute('INSERT INTO shipping(id,SONum,Customer1,address11,address12,city1,state1,zip1,country1,Customer2,address21,address22,city2,state2,zip2,country2, pid, description, qty) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)', record)
        num = num+1
    conn.commit()
    
    for SONum in SONums:
        print(SONum)
        fname = os.path.join(templateDir, 'ShippingTemplate.xlsx')
        wb = openpyxl.load_workbook(fname, data_only=True)
        shtNames = wb.get_sheet_names()
        sht = wb.get_sheet_by_name(shtNames[0])
        rows = []
        dbRows = c.execute('SELECT * FROM shipping WHERE SONum = "' + str(SONum) + '" ORDER BY id')
        for row in dbRows:
            rows.append(row)
        num = 15
        for row in rows:
            SONum = row[1]
            Customer1 = row[2]
            address11 = row[3]
            address12 = row[4]
            city1 = row[5]
            state1 = row[6]
            zip1 = row[7]
            country1 = row[8]
            Customer2 = row[9]
            address21 = row[10]
            address22 = row[11]
            city2 = row[12]
            state2 = row[13]
            zip2 = row[14]
            country2 = row[15]
            pid = row[16]
            description = row[17]
            qty = row[18]
            ###########################################
            sht["A"+str(num)].value = qty
            sht["B"+str(num)].value = pid
            sht["C"+str(num)].value = description
        
        num = 7
        if Customer1 != "":
            sht["A"+str(num)].value = Customer1
            num = num + 1
        if address11 != "":
            sht["A"+str(num)].value = address11
            num = num + 1
        if address12 != "":
            sht["A"+str(num)].value = address12
            num = num + 1
        cityStr = city1 + ', ' + state1 + " " + str(zip1)
        sht["A"+str(num)].value = cityStr
        num = num + 1
        if country1 != "":
            sht["A"+str(num)].value = country1
            num = num + 1
        num = 7
        if Customer2 != "":
            sht["D"+str(num)].value = Customer2
            num = num + 1
        if address21 != "":
            sht["D"+str(num)].value = address21
            num = num + 1
        if address22 != "":
            sht["D"+str(num)].value = address22
            num = num + 1
        cityStr = city2 + ', ' + state2 + " " + str(zip2)
        sht["D"+str(num)].value = cityStr
        num = num + 1
        if country2 != "":
            sht["D"+str(num)].value = country2
            num = num + 1            
        sht['C33'].value = SONum
        fname = os.path.join(dbDir, 'shippingSlips/' + SONum + '.xlsx')
        wb.save(fname)
        print("---------------------------------------------------------------------------------")
    
def getShipDates():
    Dates = ["03/01/2018"]
    startDate = datetime.datetime.strptime(Dates[0], "%m/%d/%Y").date()
    root = r"F:"
    records = []
    #record = [SONum1,SODate,SBDate,IVNum,CustomerName,Price,IVDate,PONum]
    for path, subdirs, files in os.walk(root):
        for name in files:
            #print(os.path.join(path, name))
            theDate = datetime.datetime.fromtimestamp(os.path.getctime(os.path.join(path, name))).date()
            if theDate >= startDate:
                try:
                    fname = os.path.join(path, name)
                    records.append(fname)
                except:
                    pass
    return records   




def autoReport(rep):
    alphaCols = ['Z','A','B','C','D','E','F','G','H','I','J','K','L']
    print("------------------------------------")
    print('Making Auto Report: ' + str(rep))
    num = 2
    fname = os.path.join(dbDir, 'BOMout.xlsx')
    wb = openpyxl.load_workbook(fname, data_only=True)
    sht = wb[rep]
    #################### CHECKS IF A DWG IS IN MATT'S STAFF FOLDER#############
    pRows = []
    rep_fields = ['id', 'pid', 'description', 'qty', 'need', 'location', 'Vendor', 'PONum', 'qtyOnOrder', 'code', 'm3', 'm6']
    rep_text = ''
    for i in range(0, len(rep_fields)):
        rep_text = rep_text + f'"{rep_fields[i]}", '
    rep_text = rep_text[:-2] + " "
    rows = c.execute('SELECT ' + rep_text + ' FROM Inventory WHERE code = "' + rep + '" and need > 0 ORDER BY (qty - need)')
    for row in rows:
        pRows.append(row)
    for row in pRows:
        for i in range(0,len(row)):
            if alphaCols[i] == 'Z':
                pass
            elif alphaCols[i] == 'E':
                val = row[3]-row[4]
                sht['E'+str(num)].value = val
                sht['F'+str(num)].value = row[i]
            elif i < 6:
                sht[alphaCols[i] + str(num)].value = row[i]
            else:
                sht[alphaCols[i+1] + str(num)].value = row[i]
        num = num + 1
    wb.save(fname)
      
def makeSepReports():
    reports = ['Machine','McMaster','Order']
    for report in reports:
        autoReport(report)
        

def makePickList(pid):
    fname = os.path.join(templateDir, 'BomPickListTemplate.xlsx')
    wb = openpyxl.load_workbook(fname, data_only=True)
    shtNames = wb.get_sheet_names()
    sht = wb.get_sheet_by_name(shtNames[0])
    num = 2
    rows = c.execute('SELECT * FROM BOM WHERE Assembly = "' + pid + '"')
    partRows = []
    theParts = []
    for row in rows:
        partRows.append(row)
    if len(partRows) > 0:
        for row in partRows:
            qty = row[5]
            rows = c.execute('SELECT * FROM Inventory WHERE pid = ?', (str(row[3]),))
            pInfo = []
            for row in rows:
                pInfo.append(row)
            for p in pInfo:
                theParts.append([pid,p[1],p[2],qty,p[3],p[5]])

    parts1 = []
    parts2 = []
    parts3 = []
    theParts = sorted(theParts, key=lambda x: x[5])
    for part in theParts:
        if part[5][:2] == "1-":
            parts1.append(part)
        elif part[5][:2] == "2-":
            parts2.append(part)
        else:
            parts3.append(part)
    allParts = [parts1,parts2,parts3]
    for part in parts1:
        part[5] = part[5][2:]
    for part in parts2:
        part[5] = part[5][2:]
    
    
    theParts.insert(0,['assembly','part','description','need','qty','location'])
    printTable2(theParts)
    
    for part in parts1:
        sht["B"+str(num)] = part[0]
        sht["C"+str(num)] = part[1]
        sht["D"+str(num)] = part[2]
        sht["E"+str(num)] = part[3]
        sht["F"+str(num)] = part[4]
        sht["G"+str(num)] = part[5]
        num = num + 1
    num = num + 2
    for part in parts2:
        sht["B"+str(num)] = part[0]
        sht["C"+str(num)] = part[1]
        sht["D"+str(num)] = part[2]
        sht["E"+str(num)] = part[3]
        sht["F"+str(num)] = part[4]
        sht["G"+str(num)] = part[5]
        num = num + 1
    num = num + 2
    for part in parts3:
        sht["B"+str(num)] = part[0]
        sht["C"+str(num)] = part[1]
        sht["D"+str(num)] = part[2]
        sht["E"+str(num)] = part[3]
        sht["F"+str(num)] = part[4]
        sht["G"+str(num)] = part[5]
        num = num + 1
    num = num + 1    
    fname = os.path.join(dbDir, 'zPickList.xlsx')
    wb.save(fname)    
    print('##########################################################################')
    print('Process finished. Please review your pick list for ' + pid)
    print('     it is saved to:      ' + fname)
    print("----------------------------------------------------------------------------")    
    

def printTable (tbl, borderHorizontal = '-', borderVertical = '|', borderCross = '+'):
    cols = [list(x) for x in zip(*tbl)]
    lengths = [max(map(len, map(str, col))) for col in cols]
    f = borderVertical + borderVertical.join(' {:>%d} ' % l for l in lengths) + borderVertical
    s = borderCross + borderCross.join(borderHorizontal * (l+2) for l in lengths) + borderCross

    print(s)
    for row in tbl:
        print(f.format(*row))
        print(s)

def printTable2 (tbl, borderHorizontal = '-', borderVertical = '|', borderCross = '+'):
    cols = [list(x) for x in zip(*tbl)]
    lengths = [max(map(len, map(str, col))) for col in cols]
    f = borderVertical + borderVertical.join(' {:%d} ' % l for l in lengths) + borderVertical
    s = borderCross + borderCross.join(borderHorizontal * (l+2) for l in lengths) + borderCross

    print(s)
    for row in tbl:
        print(f.format(*row))
        print(s)

def getAirtablePurchases():
    airtable = Airtable(base_key, 'Procurement and Fabrication', api_key)
    records = airtable.get_all(view = 'Production', formula = "FIND('Order', {Code})=1")
    airParts = []
    for record in records:
        pid = "NA"
        desc = "NA"
        PONum = "NA"
        theQty = "NA"
        status = "NA"
        result = False
        eta = "NA"
        if "Part" in record['fields'].keys():
            pid = record['fields']['Part']  
        if "Description" in record['fields'].keys():
            desc = record['fields']['Description'][:25] 
        if "PO Number" in record['fields'].keys():
            PONum = record['fields']['PO Number']  
        if "Qty" in record['fields'].keys():
            theQty = record['fields']['Qty']
        if "Status" in record['fields'].keys():
            status = record['fields']['Status']
        if "Complete" in record['fields'].keys():
            result = record['fields']['Complete']
        if "ETA" in record['fields'].keys():
            eta = record['fields']['ETA']
        
        if result:
            result = "Received"    
        else:
            if eta != "NA":
                result = str(eta)
            else:
                result = "ETA Needed"
        airParts.append([pid,desc,PONum,theQty,status,eta,result])
    return airParts
    

def updatePurchases():
    fname = os.path.join(dbDir,"onOrder.xlsx")
    wb = openpyxl.load_workbook(fname, data_only=True)
    sht = wb['Purchase Order Report']
    fname = os.path.join(dbDir,"BOMout.xlsx")
    wb2 = openpyxl.load_workbook(fname, data_only=True)
    sht2 = wb2['AirtablePurchases']
    sht3 = wb2['AirtableOrders']
    endRow = sht.max_row
    airtable = Airtable(base_key, 'Procurement and Fabrication', api_key)
    records = airtable.get_all()
    airOrders = []
    counter = 2
    check = True
    for rec in records:
        check = ""
        #print(rec['fields'].keys())
        if 'Code' in rec['fields'].keys():
            if rec['fields']['Code'] == 'Order':
                check = "a"
        if 'PO Number' in rec['fields'].keys():
            check = check + "b"
        if check == "ab":
            airOrders.append(rec)
    oldOrders = []
    newOrders = []
    sameOrders = []
    for order in airOrders:
        check = True
        PONum = order['fields']['PO Number']
        try:
            PODesc = order['fields']['Part']
        except:
            PODesc = order['fields']['Description']
        for i in range(2,endRow + 1):
            #print(PONum + " | " + sht["A"+str(i)].value)
            if PONum == sht["A"+str(i)].value:
                #print(PODesc + " | " + sht["G"+str(i)].value)
                if PODesc == sht["F"+str(i)].value:
                    #print("MATCH~!!")
                    sameOrders.append(order)
                    check = False
            #print("- - - - - - - - - - - - - - - - - - - - - - - - - - - -")
        if check:
            print(order)
            oldOrders.append(order)
    print("- - - - - - - - - - - - - - - - - - - - - - - - - - - -")
    check = True
    for i in range(2,endRow + 1):
        check = True
        PONum = sht["A"+str(i)].value
        for order in airOrders:
            if PONum == order['fields']['PO Number']:
                check = False
        if check:
            if sht["F"+str(i)].value == "METAL STOCK":
                pass
            else:
                newOrders.append([sht["F"+str(i)].value,sht["G"+str(i)].value,sht["A"+str(i)].value,sht["D"+str(i)].value,sht["K"+str(i)].value])
    print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
    for order in oldOrders:
        #MATT... FIX THIS LATER. THE QTY WAS A PROBLEM!!!!!
        try:
            print(order['fields']['PO Number'] + " | " + order['fields']['Vendor'] + " | " + order['fields']['Description'] + " | " + str(order['fields']['Qty']))
            print(order['id'])
            theRecord = airtable.match('autoIncrement', order['fields']['autoIncrement'])
            fields = {'Status': '[Remove]'}
            airtable.update(theRecord['id'], fields)
        except:
            pass
    print("================================================================================================")    
    for order in sameOrders:
        print(order['fields']['PO Number'] + " | " + order['fields']['Vendor'] + " | " + order['fields']['Description'] + " | " + str(order['fields']['Qty']))
    print("================================================================================================") 
    check = True   
    for order in newOrders:
        check = True
        print(order)
        for rec in records:
            if 'PO Number' in rec['fields'].keys():
                if rec['fields']['PO Number'] == order[2]:
                    check = False
        if check:
            sht2["C"+str(counter)].value = order[0] #ITEM ID
            sht2["D"+str(counter)].value = order[1] #DESCRIPTION
            sht2["A"+str(counter)].value = order[2] #PO NUMBER
            sht2["B"+str(counter)].value = order[3] #VENDOR
            sht2["E"+str(counter)].value = order[4] #QTY
            sht2["F"+str(counter)].value = "Order"  #CODE
            counter = counter + 1
    fname = os.path.join(dbDir,"BOMout.xlsx")

    airtable = Airtable(base_key, 'Order Status', api_key)
    air_orders = get_air_orders()
    db_orders = get_db_orders()
    air_orders.sort(key=lambda x: x[0])
    db_so_list = []
    air_so_list = []
    air_delete_list = []
    air_add_list = []
    for order in db_orders:
        db_so_list.append(order[0])
    for order in air_orders:
        air_so_list.append(order[0])
    #   now we compare airtable orders if they're in the current report
    for air_order in air_orders:
        if air_order[0] not in db_so_list:
            air_delete_list.append(air_order)
    for air_delete in air_delete_list:
        print(air_delete)
        fields = {'Status': '[Remove]'}
        airtable.update(air_delete[-1], fields)
    print("=========================================================")
    #   now we check the current report if there's any to add to airtable
    for db_order in db_orders:
        if db_order[0] not in air_so_list:
            air_add_list.append(db_order)
    for i in range(0,len(air_add_list)):
        for j in range(0,len(air_add_list[i])):
            sht3[alphaCols[j+1] + str(i+2)].value = air_add_list[i][j]

    wb2.save(fname)





        
def checkActiveBOMlist():
    makeBOMTable()
    rows = c.execute('SELECT * FROM BOM')
    bomRows = []
    for row in rows:
        bomRows.append(row)
    boms = []
    for row in bomRows:
        if row[1] not in boms:
            boms.append(row[1])
    for bom in boms:
        print(bom)


def canWeBuildIt(pid,val, lt):
    rows = c.execute('SELECT * FROM BOM WHERE Assembly = "' + pid + '"')
    partRows = []
    tableHeader = ['pid','description','have','need','diff','NEED','DIFF','location','vendor','PO','qty','code','result']
    pInfo = texttable.Texttable()
    pInfo.add_row(tableHeader)
    mInfo = [tableHeader]
    
    airParts = getAirtablePurchases()
    
    leadTime = lt
    
    
    for row in rows:
        partRows.append(row)
    if len(partRows) > 0:
        for row in partRows:
            qty = row[5]
            deltaQty = qty*val
            #rows = c.execute('SELECT * FROM Inventory WHERE pid = "' + str(row[3]) + '"')
            rows = c.execute('SELECT * FROM Inventory WHERE pid = ?', (str(row[3]),))
            for row in rows:
                part = str(row[1])
                desc = str(row[2])[:25]
                have = str(row[3])
                need = str(row[4])
                diff = row[3] - row[4]
                NEED = str(qty)
                DIFF = str(diff-deltaQty)
                loc  = str(row[5])
                vend = str(row[6])
                PONum = str(row[7])
                QTY  = str(row[8])
                code = str(row[9])
                result = "NA"
                if diff < deltaQty:
                    if code == 'Order':
                        for airPart in airParts:
                            if airPart[0] == part:
                                if airPart[6] == 'Received':
                                    result = "Received: Check inspection table and packing slips"
                                elif airPart[6] == 'ETA Needed':
                                    result = 'ETA Needed: Call the vendor and request an ETA'
                                else:
                                    result = airPart[6]
                        if result == 'NA':
                            if PONum  != "":
                                result = 'ETA Needed: Call the vendor and request an ETA'
                            else:
                                result = "NA: Look into ordering these parts ASAP"
                    elif code == 'McMaster':
                        result = 'McMaster: Check inventory/inspection table - Next Day Delivery'
                    elif code == 'Machine':
                        result = 'Machine Shop: Get on their case and get an ETA'
                                
                    
                    
                    mInfo.append([part,desc,have,need,diff,NEED,DIFF,loc,vend,PONum,QTY,code,result])
                    
                    
    print("Assembly: " + pid + "      qty: " + str(val))
    printTable(mInfo)
    


def vendorBOMout():
    f1 = os.path.join(dbDir,'PartLocations.xlsx')
    f2 = os.path.join(dbDir,'BOMout.xlsx')
    WB1 = openpyxl.load_workbook(f1,data_only=True)
    WB2 = openpyxl.load_workbook(f2,data_only=True)
    sht1 = WB1['Sheet1']
    sht2 = WB2['Order']
    m1 = sht1.max_row + 1
    m2 = sht2.max_row + 1
    
    for i in range(2,m2):
        try:
            pid = sht2['A'+str(i)].value
            v1 = "NA"
            v2 = "NA"
            for j in range(2,m1):
                pod = sht1['A'+str(j)].value
                if pid == pod:
                    v1 = sht1['K'+str(j)].value
                    v2 = sht1['L'+str(j)].value
            sht2['M'+str(i)].value = v1
            sht2['N'+str(i)].value = v2
        except:
            print(i)
    
    WB2.save(f2)
    
def Taylor():
    makeInventoryReport(False)  #True to include Airtable. False to exclude Airtable
    makeSepReports()           #You should run this if you're running the above function
    updatePurchases()
    vendorBOMout()
    print("========================================================================================")
    print("THE PROGRAM COMPLETED SUCCESSFULLY  :)")  
    
def taylor():
    makeInventoryReport(False)  #True to include Airtable. False to exclude Airtable
    makeSepReports()           #You should run this if you're running the above function
    updatePurchases()
    vendorBOMout()
    print("========================================================================================")
    print("THE PROGRAM COMPLETED SUCCESSFULLY  :)")
    
def assembler(thingy):
    makePickList(thingy)   # Work in progress...
                                        # (Assembly as it appears in SAGE, build qty, ideal lead time)


def get_air_orders():
    airtable = Airtable(base_key, 'Order Status', api_key)
    records = airtable.get_all()
    air_orders = []
    counter = 2
    for rec in records:
        air_order = []
        for air_column in air_columns:
            try:
                air_order.append(rec['fields'][air_column])
            except:
                air_order.append("")
        air_order.append(rec['id'])
        air_orders.append(air_order)
    return air_orders


def get_db_orders():
    query = "SELECT DISTINCT "
    for col in db_columns:
        query = query + col + ","
    query = query[:-1] + " FROM openOrders"
    db_orders = []
    crows = c.execute(query)
    for row in crows:
        db_orders.append(row)
    return db_orders


def update_orders():
    airtable = Airtable(base_key, 'Order Status', api_key)
    air_orders = get_air_orders()
    db_orders = get_db_orders()
    air_orders.sort(key=lambda x: x[0])
    db_so_list = []
    air_so_list = []
    air_delete_list = []
    air_add_list = []
    for order in db_orders:
        db_so_list.append(order[0])
    for order in air_orders:
        air_so_list.append(order[0])
    #   now we compare airtable orders if they're in the current report
    for air_order in air_orders:
        if air_order[0] not in db_so_list:
            air_delete_list.append(air_order)
    for air_delete in air_delete_list:
        print(air_delete)
        fields = {'Status': '[Remove]'}
        airtable.update(air_delete[-1], fields)
    print("=========================================================")
    #   now we check the current report if there's any to add to airtable
    for db_order in db_orders:
        if db_order[0] not in air_so_list:
            air_add_list.append(db_order)
    for i in range(0,len(air_add_list)):
        for j in range(0,len(air_add_list[i])):
            sht[alphaCols[i+1] + str(i+2)].value = air_add_list[i][j]



#############################################################################################################################################################
# EVERYTHING UNDER HERE
#############################################################################################################################################################


#TO TURN FUNCTION ON, REMOVE THE "#" SYMBOL FROM IN FRONT OF IT.
#TO TURN FUNCTION OFF, PUT THE "#" SYMBOL IN FRONT OF IT
#NOTE: The play button is top left green circle

# Taylor() is the default way of running the report.
Taylor()







#############################################################################################################################################################
# LEAVE AT END OF PROGRAM SO WE KNOW IF THE PROGRAM FINISHED
#############################################################################################################################################################
input('[ENTER] to exit or X-out of this window.')






