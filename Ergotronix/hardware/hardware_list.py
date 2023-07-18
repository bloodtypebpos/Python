import os
import sqlite3
import openpyxl

dbdir = r'F:\PYTHON SCRIPTS\Support Files'
os.chdir(dbdir)

fname = 'partSort.db'
conn = sqlite3.connect(fname)
c = conn.cursor()

fname = 'StockHardware.xlsx'
wb = openpyxl.load_workbook(fname)
sht = wb['Items']
pids = []
pid_rows = []
for i in range(2, sht.max_row + 1):
    pids.append(sht[f'A{i}'].value)
    row = [sht[f'A{i}'].value,
           sht[f'B{i}'].value,
           sht[f'C{i}'].value]
    pid_rows.append(row)

alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
attrs = ['QTY',
         'Thread',
         'Head',
         'Length',
         'Finish',
         'Profile',
         'Threading',
         'McMaster-Carr']

query = f'DROP TABLE IF EXISTS hw1'
c.execute(query)
conn.commit()
num = 1
query = f'CREATE TABLE IF NOT EXISTS hw1("hwid" INTEGER PRIMARY KEY, "pid" TEXT, '
for attr in attrs:
    if attr == "QTY":
        query = f'{query}"{attr}" INTEGER, '
    else:
        query = f'{query}"{attr}" TEXT, '
query = query[:-2] + ')'
c.execute(query)
conn.commit()
rows = []
for i in range(0, len(pids)):
    attr_vals = []
    attr_val = 0
    sheetname = pid_rows[i][1]
    pid_col = pid_rows[i][2]
    sht = wb[sheetname]
    for j in range(0, len(attrs)):
        try:
            for k in range(0, len(alpha)):
                if sht[f'{alpha[k]}1'].value == attrs[j]:
                    attr_vals.append(k)
        except:
            print(f'PROBLEM AT: {attrs[j]}')
    for k in range(0, len(alpha)):
        if sht[f'{alpha[k]}1'].value == pid_col:
            attr_val = alpha[k]
    print(pids[i])
    print(sheetname)
    for k in range(2, sht.max_row + 1):
        if sht[f'{attr_val}{k}'].value is not None:
            row = [num, pids[i]]
            for j in range(0, len(attrs)):
                row.append(sht[f'{alpha[attr_vals[j]]}{k}'].value)
            rows.append(row)
            num = num + 1

        else:
            pass
#    for j in range(0, len(attrs)):
#        print(f'{attrs[j]}: {alpha[attr_vals[j]]}')
    print("=======================================================================")


q = f'INSERT INTO hw1("hwid", "pid", '
for attr in attrs:
    q = f'{q}"{attr}", '
q = q[:-2] + ') VALUES('
for row in rows:
    query = f'{q}'
    for r in row:
        ro = str(r).replace('"', '``')
        query = f'{query}"{ro}", '
    query = query[:-2] + ')'
    print(row)
    c.execute(query)
conn.commit()

pid_rows = []
order_pids = []
fname = 'OpenOrdersFull.xlsx'
wb = openpyxl.load_workbook(fname)
sht = wb['Open Sales Orders']
for i in range(2, sht.max_row + 1):
    if sht[f'H{i}'].value in pids:
        if sht[f'H{i}'].value not in order_pids:
            order_pids.append(sht[f'H{i}'].value)
for pid in order_pids:
    num = 0
    for i in range(2, sht.max_row + 1):
        if sht[f'H{i}'].value == pid:
            num = num + int(sht[f'K{i}'].value)
    pid_rows.append([pid, num])
add_rows = []
for pid_row in pid_rows:
    pid = pid_row[0]
    qty = pid_row[1]
    query = F'SELECT * FROM hw1 WHERE "pid" = "{pid}"'
    crows = c.execute(query)
    print(f'{pid}:{qty}')
    rows = []
    for row in crows:
        rows.append(row)
    for row in rows:
        add_row = []
        for i in range(0, len(row)):
            if i == 2:
                val = row[i]
                val = val * qty
                add_row.append(val)
            else:
                add_row.append(row[i])
        add_rows.append(add_row)


query = f'DROP TABLE IF EXISTS hw1'
c.execute(query)
conn.commit()
num = 1
query = f'CREATE TABLE IF NOT EXISTS hw1("hwid" INTEGER PRIMARY KEY, "pid" TEXT, '
for attr in attrs:
    if attr == "QTY":
        query = f'{query}"{attr}" INTEGER, '
    else:
        query = f'{query}"{attr}" TEXT, '
query = query[:-2] + ')'
c.execute(query)
conn.commit()

q = f'INSERT INTO hw1("hwid", "pid", '
for attr in attrs:
    q = f'{q}"{attr}", '
q = q[:-2] + ') VALUES('
for row in add_rows:
    query = f'{q}'
    for r in row:
        ro = str(r).replace('"', '``')
        query = f'{query}"{ro}", '
    query = query[:-2] + ')'
    print(row)
    c.execute(query)
conn.commit()

query = 'SELECT DISTINCT "McMaster-Carr" FROM hw1'
crows = c.execute(query)
mcmasters = []
for row in crows:
    mcmasters.append(row[0])

add_rows = []

for mcmaster in mcmasters:
    if mcmaster == 'None':
        pass
    else:
        rows = []
        # print(mcmaster)
        query = f'SELECT * FROM hw1 WHERE "McMaster-Carr" = "{mcmaster}"'
        crows = c.execute(query)
        for row in crows:
            rows.append(row)
        qty = 0
        add_row = []
        for row in rows:
            qty = qty + row[2]
        row = rows[0]
        for i in range(0, len(row)):
            if i == 2:
                add_row.append(qty)
            else:
                add_row.append(row[i])
        add_rows.append(add_row)

print("================================================================")
print("================================================================")
print("================================================================")
num = 2
fname = 'zPickList.xlsx'
wb = openpyxl.load_workbook(fname)
sht = wb['Sheet2']
sht[f'A1'].value = 'LINE'
for i in range(0, len(attrs)):
    sht[f'{alpha[i+1]}1'].value = attrs[i]
for row in add_rows:
    print(row)
    sht[f'A{num}'].value = num-1
    val = 1
    for i in range(2, len(row)):
        sht[f'{alpha[val]}{num}'].value = row[i]
        val = val + 1
    num = num + 1
wb.save(fname)






